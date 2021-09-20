from datetime import date
import pathlib
import csv
from getmac import get_mac_address as gma

import pandas as pd

import xlsxwriter

import openpyxl
from openpyxl import load_workbook

from file_grabber import grab_files

cwd = pathlib.Path.cwd()
print(gma())


grab_files(
    [
        "03_4_PS_Enroll.csv",
        "03_5_PS_GradeProg.csv",
        "06_5A_Public_PreK_Stu_Link.csv",
        "racedata.csv",
    ],
    cwd,
)

## -------------------------------------------------------------------
## load enroll files
df_raw_enroll = pd.read_csv(cwd / "data_files" / "03_4_PS_Enroll.csv", dtype=str)

# drop all non-essential columns from PS_Enroll file
df_enroll = df_raw_enroll.iloc[:, [1, 2, 3, 4, 5, 6, 7, 8, 12, 17]].copy()

# drop students that are no longer enrolled
df_enroll = df_enroll.loc[df_enroll["ENRENDDATE"].isnull()]

## -------------------------------------------------------------------
## load gradeprog file
df_grades = pd.read_csv(cwd / "data_files" / "03_5_PS_GradeProg.csv", dtype=str)

# drop all non-essential columns from PS_Enroll file
df_grades = df_grades.iloc[:, [2, 3]].copy()

# combine df's into one df
enroll = pd.merge(df_enroll, df_grades, how="left", on=["PERMNUMBER"])

# remove the unused column
enroll = enroll.drop("ENRENDDATE", 1)

# count EE
ee_kids = enroll[enroll["GRADE"].isin(["EE"])]

# remove EEE students from enrollment file
enroll = enroll[~enroll["GRADE"].isin(["EE"])]

# save number of EE kids for later reporting
ee_kids_num = len(ee_kids)

## -------------------------------------------------------------------
## load PK info files
df_PK = pd.read_csv(cwd / "data_files" / "06_5A_Public_PreK_Stu_Link.csv", dtype=str)

# get number of ALL PK
all_pk = len(df_PK)

# new df with only PK kids from three schools
df_PK = df_PK.loc[
    (df_PK["PKID"] == "PK00302")
    | (df_PK["PKID"] == "PK00288")
    | (df_PK["PKID"] == "PK00201")
]

# drop all non-essential columns from PK file
df_PK = df_PK.iloc[:, [2, 3]].copy()

# calculate act 166 kiddos
act_166_num = all_pk - len(df_PK)

# combine enroll and PK df's into one df
final_enroll = pd.merge(enroll, df_PK, how="left", on=["PERMNUMBER"])

# exclude home daycares from final list
total_enroll = final_enroll.loc[
    ~((final_enroll["GRADE"] == "PK") & (final_enroll["PKID"].isnull())), :
]

## -------------------------------------------------------------------
## get total numbers by school and grade

# new df by school and grade
summary = total_enroll.groupby(["ENRORGID", "GRADE"]).size()

# succint df of numbers by school and grade
grid = total_enroll.groupby(["ENRORGID", "GRADE"]).size().unstack()

# transpose df so they are easier to read
grid = grid.T

# add row totals
grid["Total"] = grid.sum(axis=1)

# add column totals
grid.loc["Total"] = grid.sum(numeric_only=True, axis=0)

# take out Pk and Kinder to add to top
pre = grid.loc[["PK"]]
kinder = grid.loc[["KF"]]

# drop PK and Kinder
grid = grid.drop(["KF", "PK"])

# add PK and kind back in at top
final_numbers = pd.concat([kinder, grid])
final_numbers = pd.concat([pre, final_numbers])

# rename columns for report
final_numbers.rename(
    columns={"PS115": "FCS", "PS142": "HES", "PS187": "MVU", "PS295": "SWA"},
    inplace=True,
)

# move MVU to the end
mvu = final_numbers.pop("MVU")
final_numbers.insert(3, "MVU", mvu)

## -------------------------------------------------------------------
## Begin racial report (including multi-racial)

# fresh copy for new df
multi = total_enroll.copy()

# setup columns as numbers
multi["ETHNO"].replace("2", 0, inplace=True)
multi["RACE_AMI"].replace("Y", 1, inplace=True)
multi["RACE_AMI"].replace("N", 0, inplace=True)
multi["RACE_ASI"].replace("Y", 1, inplace=True)
multi["RACE_ASI"].replace("N", 0, inplace=True)
multi["RACE_AFA"].replace("Y", 1, inplace=True)
multi["RACE_AFA"].replace("N", 0, inplace=True)
multi["RACE_NAT"].replace("Y", 1, inplace=True)
multi["RACE_NAT"].replace("N", 0, inplace=True)
multi["RACE_WHT"].replace("Y", 1, inplace=True)
multi["RACE_WHT"].replace("N", 0, inplace=True)

# ensure they are integers
multi["ETHNO"] = multi["ETHNO"].astype(int)
multi["RACE_AMI"] = multi["RACE_AMI"].astype(int)
multi["RACE_ASI"] = multi["RACE_ASI"].astype(int)
multi["RACE_AFA"] = multi["RACE_AFA"].astype(int)
multi["RACE_NAT"] = multi["RACE_NAT"].astype(int)
multi["RACE_WHT"] = multi["RACE_WHT"].astype(int)

# new row summing up the columns
multi["multi"] = multi.apply(
    lambda row: row.RACE_AMI
    + row.RACE_ASI
    + row.RACE_AFA
    + row.RACE_NAT
    + row.RACE_WHT,
    axis=1,
)

# remove value (make 0) from individual columns if multi is greater than 1
multi.loc[
    multi.multi > 1, ["RACE_AMI", "RACE_ASI", "RACE_AFA", "RACE_NAT", "RACE_WHT"]
] = (0, 0, 0, 0, 0)

# remove value (make 0) from individual columns if ethno is hispanic
multi.loc[
    multi.ETHNO == 1,
    ["RACE_AMI", "RACE_ASI", "RACE_AFA", "RACE_NAT", "RACE_WHT", "multi"],
] = (0, 0, 0, 0, 0, 0)

# if multi = 1, then replace with 0 meaning not multiracial
multi["multi"].replace(1, 0, inplace=True)

# if multi greater than 1, replace with one for multiracial
multi.loc[multi.multi > 1, "multi"] = 1

# drop all non-essential columns from PS_Enroll file
multi = multi.iloc[:, [0, 2, 3, 4, 5, 6, 7, 11]].copy()

# find and replace school numbers with names
multi["ENRORGID"].replace("PS115", "FCS", inplace=True)
multi["ENRORGID"].replace("PS142", "HES", inplace=True)
multi["ENRORGID"].replace("PS187", "MVU", inplace=True)
multi["ENRORGID"].replace("PS295", "SWA", inplace=True)

# # replace column names
# multi.rename(
#     columns={
#         0: "Hispanic or Latino",
#         1: "American Indian or Alaska Native",
#         2: "Asian",
#         3: "Black or African American",
#         4: "Native Hawaiian or Other Pacific Islander",
#         5: "White",
#         6: "Multiracial",
#     },
#     inplace=True,
# )

# new grid with all data
multi_race_data = multi.groupby(["ENRORGID"]).sum()

# replace column names
multi_race_data.rename(
    columns={
        "ETHNO": "Hispanic or Latino",
        "RACE_AMI": "American Indian or Alaska Native",
        "RACE_ASI": "Asian",
        "RACE_AFA": "Black or African American",
        "RACE_NAT": "Native Hawaiian or Other Pacific Islander",
        "RACE_WHT": "White",
        "multi": "Multiracial",
    },
    inplace=True,
)

# add totals to rows, then columns
multi_race_data["Total"] = multi_race_data.sum(axis=1)
multi_race_data.loc["Total"] = multi_race_data.sum(numeric_only=True, axis=0)

## -------------------------------------------------------------------
## Begin free and reduced

# grab fresh copy of df
free = total_enroll.copy()

# get just school, grade, and free/reduced info
free = free.iloc[:, [0, 8, 9]].copy()

# rename columns for report
free.rename(
    columns={"NSLELG": "Free/Reduced", "ENRORGID": "School", "GRADE": "Grade"},
    inplace=True,
)

# find and replace values
free["Free/Reduced"].replace("96", 0, inplace=True)
free["Free/Reduced"].replace("01", 1, inplace=True)
free["Free/Reduced"].replace("02", 1, inplace=True)
free["School"].replace("PS115", "FCS", inplace=True)
free["School"].replace("PS142", "HES", inplace=True)
free["School"].replace("PS187", "MVU", inplace=True)
free["School"].replace("PS295", "SWA", inplace=True)

# new grid with totals of values
grid3 = free.groupby(["School", "Grade"]).sum().unstack()

# add totals to rows
grid3["F/R Total"] = grid3.sum(axis=1)

# add total enrollment column
grid3["Total Enrollment"] = free["School"].value_counts()

# add totals for each grade in final row
grid3.loc["Total"] = grid3.sum(numeric_only=True, axis=0)

# add percentage column
grid3["F/R Percentage"] = grid3["F/R Total"] / grid3["Total Enrollment"]
grid3["F/R Percentage"] = grid3["F/R Percentage"].astype(float).map("{:.2%}".format)

## ------------------------------------------------------------------------------
## begin single race data

# grab fresh copy of enrollment
race = total_enroll.copy()


# load race data from PS export
df_raw_race = pd.read_csv(cwd / "data_files" / "racedata.csv", header=None, dtype=str)
df_raw_race.columns = ["PERMNUMBER", "Race"]

# merge data to enrollment
race_data = pd.merge(race, df_raw_race, how="left", on=["PERMNUMBER"])


# reduce data to only school and race column
race_data = race_data.iloc[:, [0, 1, 11]].copy()

# save to csv for old school python processing
race_data.to_csv(cwd / "data_files" / "race_data.csv", index=False)

race_file = cwd / "data_files" / "race_data.csv"

# create variable for totals
mvu_w = 0
mvu_i = 0
mvu_a = 0
mvu_b = 0
mvu_h = 0
mvu_p = 0

fcs_w = 0
fcs_i = 0
fcs_a = 0
fcs_b = 0
fcs_h = 0
fcs_p = 0

hes_w = 0
hes_i = 0
hes_a = 0
hes_b = 0
hes_h = 0
hes_p = 0

swa_w = 0
swa_i = 0
swa_a = 0
swa_b = 0
swa_h = 0
swa_p = 0

with open(race_file, "r") as data:
    for line in csv.reader(data):
        if line[1]:
            # print(line)
            if line[0] == "PS187":
                if line[2] == "W":
                    mvu_w += 1
                elif line[2] == "I":
                    mvu_i += 1
                elif line[2] == "A":
                    mvu_a += 1
                elif line[2] == "B":
                    mvu_b += 1
                elif line[2] == "H":
                    mvu_h += 1
                elif line[2] == "P":
                    mvu_p += 1
            if line[0] == "PS115":
                if line[2] == "W":
                    fcs_w += 1
                elif line[2] == "I":
                    fcs_i += 1
                elif line[2] == "A":
                    fcs_a += 1
                elif line[2] == "B":
                    fcs_b += 1
                elif line[2] == "H":
                    fcs_h += 1
                elif line[2] == "P":
                    fcs_p += 1
            if line[0] == "PS142":
                if line[2] == "W":
                    hes_w += 1
                elif line[2] == "I":
                    hes_i += 1
                elif line[2] == "A":
                    hes_a += 1
                elif line[2] == "B":
                    hes_b += 1
                elif line[2] == "H":
                    hes_h += 1
                elif line[2] == "P":
                    hes_p += 1
            if line[0] == "PS295":
                if line[2] == "W":
                    swa_w += 1
                elif line[2] == "I":
                    swa_i += 1
                elif line[2] == "A":
                    swa_a += 1
                elif line[2] == "B":
                    swa_b += 1
                elif line[2] == "H":
                    swa_h += 1
                elif line[2] == "P":
                    swa_p += 1


fcs_total = fcs_h + fcs_i + fcs_a + fcs_b + fcs_p + fcs_w
hes_total = hes_h + hes_i + hes_a + hes_b + hes_p + hes_w
swa_total = swa_h + swa_i + swa_a + swa_b + swa_p + swa_w
mvu_total = mvu_h + mvu_i + mvu_a + mvu_b + mvu_p + mvu_w

df_race = pd.DataFrame(
    [
        (
            fcs_h,
            fcs_i,
            fcs_a,
            fcs_b,
            fcs_p,
            fcs_w,
            fcs_total,
        ),
        (
            hes_h,
            hes_i,
            hes_a,
            hes_b,
            hes_p,
            hes_w,
            hes_total,
        ),
        (
            swa_h,
            swa_i,
            swa_a,
            swa_b,
            swa_p,
            swa_w,
            swa_total,
        ),
        (
            mvu_h,
            mvu_i,
            mvu_a,
            mvu_b,
            mvu_p,
            mvu_w,
            mvu_total,
        ),
        (
            mvu_h + fcs_h + hes_h + swa_h,
            mvu_i + fcs_i + hes_i + swa_i,
            mvu_a + fcs_a + hes_a + swa_a,
            mvu_b + fcs_b + hes_b + swa_b,
            mvu_p + fcs_p + hes_p + swa_p,
            mvu_w + fcs_w + hes_w + swa_w,
            mvu_total + fcs_total + hes_total + swa_total,
        ),
    ],
    index=["FCS", "HES", "SWA", "MVU", "Total"],
    columns=(
        "Hispanic",
        "American Indian",
        "Asian",
        "Black",
        "Islander",
        "White",
        "Total",
    ),
)

# print(df)

## ------------------------------------------------------------------------------
## begin presenting data in excell sheet

# get datetime for naming
today = date.today()
d1 = today.strftime("%d/%m/%Y")
d1 = d1.replace("/", "-")

# create string for data numbers name
data_sheet_name = "Student_Data_Numbers_" + d1 + ".xlsx"

# create excel sheet with data
writer = pd.ExcelWriter(cwd / "complete_data" / data_sheet_name, engine="xlsxwriter")


final_numbers.to_excel(writer, sheet_name="Students by Grade and School")
df_race.to_excel(writer, sheet_name="Students by Single Ethnicity")
multi_race_data.to_excel(writer, sheet_name="Students by Ethnic - wMulti")
grid3.to_excel(writer, sheet_name="Students by F-R Lunch")

writer.save()

# add 166 data to first sheet
# wb = load_workbook(filename="Student Data Numbers " + d1 + ".xlsx")
wb = load_workbook(filename=cwd / "complete_data" / data_sheet_name)
fp = wb["Students by Grade and School"]
fp["H1"] = "EEE:"
fp["I1"] = ee_kids_num
fp["H2"] = "Offsite 166:"
fp["I2"] = act_166_num
wb.save(cwd / "complete_data" / data_sheet_name)

# /Users/rgregory/Documents/trial

my_file = cwd / "complete_data" / data_sheet_name
# to_file = "/Users/rgregory/Documents/trial/" + data_sheet_name
to_file = "/Users/admin/Documents/student_numbers/" + data_sheet_name
my_file.rename(to_file)
